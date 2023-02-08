import pandas as pd
import os
import openpyxl
import glob
import itertools
import csv


def  course_check(file_1,file_2):
    num1 = file_1.rstrip(".xyz").lstrip("line")
    num2 = file_2.rstrip(".xyz").lstrip("line")
    df1 = pd.read_table(file_1, names=['x', 'y', 'z1'], delim_whitespace=True)
    df2 = pd.read_table(file_2, names=['x', 'y', 'z2'], delim_whitespace=True)
    print('df1:' + str(len(df1['x'].to_list())))
    print('df2:' + str(len(df2['x'].to_list())))
    df = df1.merge(df2, on=['x', 'y'], how='inner')
    df['dif'] = df['z2'] - df['z1']
    count = len(df['dif'].to_list())
    print('df1+2:'+ str(count))
    #df.to_csv()
    #print(df)
    df_result = df.query('-0.1 < = dif < = 0.1')
    df_result = df_result.sort_values(['x', 'y'])
    a = df_result.head(20)
    b = df_result.tail(20)
    print(a)
    print(b)
    name = num1 + '_' +num2
    y1 = a.iloc[7][0] #1点目　y座標　メッシュ
    y2 = b.iloc[12][0] #2点目　y座標　メッシュ
    x1 = a.iloc[7][1]#1点目　x座標　メッシュ
    x2 = b.iloc[12][1]#2点目　x座標　メッシュ
    z1_1 = a.iloc[7][2]#1点目　標高値　メッシュ（z1)
    z1_2 = a.iloc[7][3]#1点目　標高値　メッシュ（z2)
    z1_kousa = a.iloc[7][4]#1点目　標高値　較差（z1)
    z2_1 = b.iloc[12][2]#2点目　標高値　メッシュ（z1)
    z2_2 = b.iloc[12][3]#2点目　標高値　メッシュ（z1)
    z3_kousa = b.iloc[12][4]#2点目　標高値　較差（z1)
    return name,y1,y2,x1,x2,z1_1,z1_2,z1_kousa,z2_1,z2_2,z3_kousa,count

def to_csv(dir):
    os.chdir(os.path.join(dir, "5_line"))
    files = glob.glob("*.xyz")
    files.sort()
    #print(files)
    conbi = []
    for x in itertools.combinations(files, 2):
        conbi.append(x)
    add_num = 0
    name = []
    name1 = []
    name2 = []
    y1 = [] # 1点目　y座標　メッシュ
    y2 = []  # 2点目　y座標　メッシュ
    x1 = []  # 1点目　x座標　メッシュ
    x2 = []  # 2点目　x座標　メッシュ
    z1_1 = []  # 1点目　標高値　メッシュ（z1)
    z1_2 = []  # 1点目　標高値　メッシュ（z2)
    z1_kousa = []  # 1点目　標高値　較差（z1)
    z2_1 = []  # 2点目　標高値　メッシュ（z1)
    z2_2 = [] # 2点目　標高値　メッシュ（z1)
    z2_kousa = []  # 2点目　標高値　較差（z1)
    count = []
    for i in conbi:
        text = str(i)
        text = text.replace('(', '')
        text = text.replace("'", '')
        text = text.replace(')', '')
        text = text.replace(' ', '')
        print(text)
        lst = text.split(',')
        file1 = str(lst[0])
        file2 = str(lst[1])
        print(lst)
        try:
            ans = course_check(file1, file2)
            name.append(ans[0])
            name1.append(file1)
            name2.append(file2.split('-')[0])
            y1.append(ans[1])  # 1点目　y座標　メッシュ
            y2.append(ans[2])   # 2点目　y座標　メッシュ
            x1.append(ans[3]) # 1点目　x座標　メッシュ
            x2.append(ans[4])  # 2点目　x座標　メッシュ
            z1_1.append(ans[5]) # 1点目　標高値　メッシュ（z1)
            z1_2.append(ans[6])  # 1点目　標高値　メッシュ（z2)
            z1_kousa.append(ans[7])  # 1点目　標高値　較差（z1)
            z2_1.append(ans[8])  # 2点目　標高値　メッシュ（z1)
            z2_2.append(ans[9])  # 2点目　標高値　メッシュ（z1)
            z2_kousa.append(ans[10])  # 2点目　標高値　較差（z1)
            count.append(ans[11])
        except IndexError:
            pass
    df_result = pd.DataFrame({'name': name,'name1': name1 ,'name2': name2 ,'y1': y1, 'x1': x1,'z1_1':z1_1,'z1_2':z1_2,'z1_kousa':z1_kousa,'y2': y2, 'x2': x2,'z2_1':z2_1,'z2_2':z2_2,'z2_kousa':z2_kousa,'mesh_count':count})
    df_result.to_csv("test.csv")
    f_list = df_result['name2'].to_list()
    course = df_result['name1'].to_list()
    f_list = list(set(f_list))
    return lst, course,f_list

def csv_filter(path,txt,course,f_list):
    df = pd.read_csv(txt, header=0)
    print(df['mesh_count'].idxmax())
    df_list = []
    for i in course:
        for f in f_list:
            df_and = df[(df['name1'] == i) & ~(df['name2'] == f)]
            try:
                max_row = df_and['mesh_count'].idxmax()
                query = df_and.loc[[max_row]]
                df_list.append(query)
                print(max_row)
            except (IndexError, ValueError):
                pass

    df_result = pd.concat(df_list)
    df_result.to_csv(os.path.join(path, r'5_line\before_filter.csv'))
    df_result = df_result.drop_duplicates()
    print(df_result)
    df_result.drop('Unnamed: 0', axis=1)
    df_result.to_csv(os.path.join(path, r'5_line\filter.csv'))
    return df_result

def to_excel(path,df,add_num):
    ex_path = os.path.join(path, "3_公共_精度管理表1式") + "\様式8_コース間検証精度管理表.xlsx"
    wb = openpyxl.load_workbook(ex_path)
    # df_result = pd.DataFrame(['name','name1','name2' ,'y1', 'x1','z1_1','z1_2','z1_kousa','y2', 'x2','z2_1','z2_2','z2_kousa','mesh_count'])
    name = df.name.values.tolist()
    y1 = df.y1.values.tolist()   # 1点目　y座標　メッシュ
    y2 = df.y2.values.tolist()  # 2点目　y座標　メッシュ
    x1 = df.x1.values.tolist()  # 1点目　x座標　メッシュ
    x2 = df.x2.values.tolist() # 2点目　x座標　メッシュ
    z1_1 = df.z1_1.values.tolist()  # 1点目　標高値　メッシュ（z1)
    z1_2 = df.z1_2.values.tolist()  # 1点目　標高値　メッシュ（z2)
    z1_kousa = df.z1_kousa.values.tolist()  # 1点目　標高値　較差（z1)
    z2_1 = df.z2_1.values.tolist()  # 2点目　標高値　メッシュ（z1)
    z2_2 = df.z2_2.values.tolist()  # 2点目　標高値　メッシュ（z1)
    z2_kousa = df.z2_kousa.values.tolist()  # 2点目　標高値　較差（z1)
    for i in range(len(name)):
        ws = wb.copy_worksheet(wb['様式8'])
        sheets = wb.sheetnames
        sheet_num = sheets.index(ws.title)
        ws.title = name[i]
        ws['R8'] = 'C-' + name[i].split('_')[0]
        ws['X8'] = 'C-' + name[i].split('_')[1]
        if sheet_num == 1:
            ws['B9'] = str(sheet_num + add_num)
            ws['B11'] = str(sheet_num + 1 + add_num)
        else:
            ws['B9'] = str(sheet_num * 2 - 1 + add_num)
            ws['B11'] = str(sheet_num * 2 + add_num)
        ws['L9'] = y1[i]#y1
        ws['L11'] = y2[i]#y2
        ws['F9'] = x1[i]#x1
        ws['F11'] = x2[i]#x2
        ws['R9'] = z1_1[i]
        ws['x9'] = z1_2[i]
        ws['ad9'] = z1_kousa[i]
        ws['R11'] = z2_1[i]
        ws['x11'] = z2_1[i]
        ws['ad11'] = z2_kousa[i]
    wb.save(ex_path)

def main(path):
    lst = to_csv(path)
    txt = os.path.join(path, r'5_line\test.csv')
    df_result = csv_filter(path, txt, lst[1], lst[2])
    add_num = 0
    to_excel(path, df_result, add_num)
    return

if __name__ == '__main__':
    path = r'H:\共有ドライブ\221018_TLcloud\01_CLOUD\TLO\県北土地開発\221222_一般国道251号道路改良工事(測量業務委託その6)\解析'
    main(path)








